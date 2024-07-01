import openpyxl
from datetime import datetime
import os

def check_engeneer() -> tuple[str]:
    engeneer_1, engeneer_2 = None, None
    wookbook = openpyxl.load_workbook("./данные/табель_сотрудников.xlsx")
    worksheet = wookbook.active
    now_date = int(datetime.now().strftime('%d'))
    for column in worksheet.iter_cols(): 
        column_name = column[1].value 
        if column_name == now_date: 
            for item, cell in enumerate(column): 
                if item == 0 or item == 1: 
                    continue
                elif cell.value == 1:
                    engeneer_1 = worksheet[f"A{item+1}"].value
                elif cell.value == 2:  
                    engeneer_2 = worksheet[f"A{item+1}"].value
    return (engeneer_1, engeneer_2)

def check_engeneer_info() -> str:
    engeneer = None
    wookbook = openpyxl.load_workbook("./данные/табель_сотрудников.xlsx")
    worksheet = wookbook.active
    now_date = datetime.now().strftime('%d %H').split()
    if (int(now_date[1]) >= 20) or (8 <= int(now_date[1]) < 20):
        for column in worksheet.iter_cols():
            column_name = column[1].value
            if column_name == int(now_date[0]):
                for item, cell in enumerate(column):
                    if item == 0 or item == 1:
                        continue
                    elif 8 <= int(now_date[1]) < 20 and cell.value == 1:
                        engeneer = worksheet[f"A{item+1}"].value
                    elif int(now_date[1]) >= 20 and cell.value == 2:
                        engeneer = worksheet[f"A{item+1}"].value
    elif int(now_date[1]) < 8:
        try:
            res = int(now_date[0]) - 1
        except Exception:
            for column in worksheet.iter_cols():
                column_name = column[1].value
                if str(column_name) == '30/31':
                    for item, cell in enumerate(column):
                        if item == 0 or item == 1:
                            continue
                        elif cell.value == 2:
                            engeneer = worksheet[f"A{item+1}"].value
        else:
            for column in worksheet.iter_cols():
                column_name = column[1].value
                if column_name == (int(now_date[0]) - 1):
                    for item, cell in enumerate(column):
                        if item == 0 or item == 1:
                            continue
                        elif cell.value == 2:
                            engeneer = worksheet[f"A{item+1}"].value
    return (engeneer)

def check_engeneer_info_start() -> str:
    engeneer = None
    wookbook = openpyxl.load_workbook("./данные/табель_сотрудников.xlsx")
    worksheet = wookbook.active
    filename = "./данные/изменения_в_работе.txt"
    res_date = os.path.getmtime(filename)
    now_date = datetime.fromtimestamp(res_date).strftime("%d %H").split()
    if (int(now_date[1]) >= 20) or (8 <= int(now_date[1]) < 20):
        for column in worksheet.iter_cols():
            column_name = column[1].value
            if column_name == int(now_date[0]):
                for item, cell in enumerate(column):
                    if item == 0 or item == 1:
                        continue
                    elif 8 <= int(now_date[1]) < 20 and cell.value == 1:
                        engeneer = worksheet[f"A{item+1}"].value
                    elif int(now_date[1]) >= 20 and cell.value == 2:
                        engeneer = worksheet[f"A{item+1}"].value
    elif int(now_date[1]) < 8:
        try:
            res = int(now_date[0]) - 1
        except Exception:
            for column in worksheet.iter_cols():
                column_name = column[1].value
                if str(column_name) == '30/31':
                    for item, cell in enumerate(column):
                        if item == 0 or item == 1:
                            continue
                        elif cell.value == 2:
                            engeneer = worksheet[f"A{item+1}"].value
        else:
            for column in worksheet.iter_cols():
                column_name = column[1].value
                if column_name == (int(now_date[0]) - 1):
                    for item, cell in enumerate(column):
                        if item == 0 or item == 1:
                            continue
                        elif cell.value == 2:
                            engeneer = worksheet[f"A{item+1}"].value
    return (engeneer)