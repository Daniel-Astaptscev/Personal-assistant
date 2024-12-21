# ///////////////////////////////////////////////////////////////
# IMPORT 
# ///////////////////////////////////////////////////////////////
from bs4 import BeautifulSoup
import openpyxl
import requests

# ///////////////////////////////////////////////////////////////
# Variables
# ///////////////////////////////////////////////////////////////
result_max, result_min, result_average, answer = [], [], [], []
index_month = {1: "январь", 2: "февраль", 3: "март", 4: "апрель",
               9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"}
month_xlsx = {"январь": 2, "февраль": 6, "март": 10, "апрель": 14,
              "сентябрь": 18, "октябрь": 22, "ноябрь": 26, "декабрь": 30}


# ///////////////////////////////////////////////////////////////
# The work of the program
# ///////////////////////////////////////////////////////////////
def calculate_temperature(item: str, year: int, month: int,
                          fact_value: float) -> list[str]:
    if len(answer) > 0:
        answer.clear()

    # Get a response from the server
    # ///////////////////////////////////////////////////////////////
    def check_url(url: str) -> None:
        if url.status_code == 200:
            answer.append(f'запрос к сайту по пути - {item} был выполнен!\n')
        else:
            answer.append(
                'программа обнаружила исключение: отсутствует доступ к серверу...\n')

    # Algorithm
    # ///////////////////////////////////////////////////////////////
    def average_value(lst_max: list[str], lst_min: list[str]) -> None:
        clear_temp = [[value.text for value in lst_max],
                      [value.text for value in lst_min]]
        temp = [[int(value.replace('°', '')) for value in clear_temp[0]],
                [int(value.replace('°', '')) for value in clear_temp[1]]]

        try:
            total = round(sum(temp[0]) / len(temp[0]), 1)
            result_max.append(total)
            result_average.append(round((total - fact_value), 1))
            total = round(sum(temp[1]) / len(temp[1]), 1)
            result_min.append(total)
        except Exception:
            answer.append(
                f'обнаружено исключение - полученные значения в запросе не могут быть обработаны...\n')
        else:
            answer.append(
                'произведена выгрузка собранных данных - сохранение в список: успешно!\n')

    # Site parsing
    # ///////////////////////////////////////////////////////////////
    try:
        month = index_month[month]
        url = f'https://{item}.nuipogoda.ru/{month}-{year}'
        page = requests.get(url)
        check_url(page)
        soup = BeautifulSoup(page.text, "html.parser")
    except Exception:
        answer.append(f'обнаружено исключение - запрос к серверу невозможен...\n')
    average_value(soup.find_all('span', class_='max'),
                  soup.find_all('span', class_='min'))
    return answer

# Saving values
# ///////////////////////////////////////////////////////////////
def saving_temperature(year: int, month: int, fact_value: float) -> bool:
    try:
        temp_xlsx = openpyxl.load_workbook(
            './данные/среднемесячная_температура_по_годам.xlsx')
        sheet = temp_xlsx[str(year)]
        month = index_month[month]
        for num_row in range(3, 27):
            sheet.cell(row=num_row, column=month_xlsx[month]).value = result_max[num_row - 3]
            sheet.cell(row=num_row, column=month_xlsx[month] + 1).value = result_min[num_row - 3]
            sheet.cell(row=num_row, column=month_xlsx[month] + 2).value = fact_value
            sheet.cell(row=num_row, column=month_xlsx[month] + 3).value = result_average[num_row - 3]
        temp_xlsx.save('./данные/среднемесячная_температура_по_годам.xlsx')
        result_max.clear()
        result_min.clear()
        result_average.clear()
        return True
    except Exception:
        return False
